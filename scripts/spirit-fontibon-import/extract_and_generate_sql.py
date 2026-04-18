"""
Extractor y normalizador de data FONTIBON 2026.xlsx
Genera un script SQL transaccional para Supabase.

Uso:
    python extract_and_generate_sql.py
"""

import openpyxl
from datetime import date, datetime
import re
import sys
import io

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

EXCEL_PATH = 'C:/Users/Usuario/Documents/DOCUMENTACION SPORTMPAS/FONTIBON 2026.xlsx'
OUTPUT_SQL = 'c:/Users/Usuario/Documents/demo/sportmaps-demo/scripts/spirit-fontibon-import/01_create_test_school_fontibon.sql'
OWNER_EMAIL = 'jreyes@gmail.com'
SCHOOL_NAME = 'Spirit Fontibon (Test)'
BRANCH_NAME = 'Fontibon'
PROGRAM_NAME = 'Cheerleading All-Star Butterfly'

# IDs hardcoded (confirmados desde Supabase) - escuela y sede de jreyes ya existentes
SCHOOL_ID = 'aaaaaaaa-bbbb-cccc-dddd-eeeeeeeeeeee'
BRANCH_ID = 'ffffffff-1111-2222-3333-444444444444'

# Fixes manuales confirmados por el usuario
DATE_OVERRIDES = {
    ('BUTTERFLY', 'Hellen Samantha Padilla Caceres'): date(2018, 8, 30),
    ('SPRINKLES', 'Luciana Bermudez sanchez'): date(2020, 6, 16),
    ('BOMBSHELLS', 'Isabella Ardila Ramirez'): date(2011, 12, 20),
    ('BOMBSQUAD', 'Anne Silvana Alarcon'): date(2014, 2, 24),
}

# Mapeo de columnas por hoja (columna -> tipo)
# tipos: name, doc_type, doc_number, dob, age, tshirt, eps, phone_athlete, parent_name, phone_parent
SHEET_COLUMN_MAP = {
    'SPRINKLES':  {1: 'name', 2: 'doc_type', 3: 'doc_number', 4: 'age', 5: 'dob', 6: 'tshirt', 7: 'eps', 8: 'phone_athlete', 9: 'parent_name', 10: 'phone_parent'},
    'BUTTERFLY':  {1: 'name', 2: 'doc_type', 3: 'doc_number', 4: 'dob', 5: 'age', 6: 'eps', 7: 'parent_name', 8: 'phone_parent', 9: 'tshirt', 10: 'phone_parent2'},
    'BOMBSQUAD':  {1: 'name', 2: 'doc_type', 3: 'doc_number', 4: 'dob', 5: 'tshirt', 6: 'eps', 7: 'parent_name', 8: 'phone_parent'},
    'LEGENDS ':   {1: 'name', 2: 'doc_type', 3: 'doc_number', 4: 'dob', 5: 'tshirt', 6: 'eps', 7: 'phone_athlete', 8: 'parent_name', 9: 'phone_parent'},
    'FIRESQUAD':  {1: 'name', 2: 'doc_type', 3: 'doc_number', 4: 'dob', 5: 'tshirt', 6: 'eps', 7: 'phone_athlete', 8: 'parent_name', 9: 'phone_parent'},
    'BOMBSHELLS': {1: 'name', 2: 'doc_type', 3: 'doc_number', 4: 'dob', 5: 'tshirt', 6: 'eps', 7: 'parent_name', 8: 'phone_parent'},
}

TEAM_UUIDS = {
    'SPRINKLES':  "'11111111-1111-1111-1111-000000000001'",
    'BUTTERFLY':  "'11111111-1111-1111-1111-000000000002'",
    'BOMBSQUAD':  "'11111111-1111-1111-1111-000000000003'",
    'LEGENDS':    "'11111111-1111-1111-1111-000000000004'",
    'FIRESQUAD':  "'11111111-1111-1111-1111-000000000005'",
    'BOMBSHELLS': "'11111111-1111-1111-1111-000000000006'",
}

TEAM_AGE_GROUPS = {
    'SPRINKLES': 'Mini (5-8)',
    'BUTTERFLY': 'Mini (4-10)',
    'BOMBSQUAD': 'Youth (10-17)',
    'LEGENDS': 'Senior (12-21)',
    'FIRESQUAD': 'Youth (6-13)',
    'BOMBSHELLS': 'Junior (8-16)',
}


def strip_accents(s):
    """Quita acentos simples para matching."""
    if not s:
        return ''
    replacements = str.maketrans('áéíóúÁÉÍÓÚñÑ', 'aeiouAEIOUnN')
    return s.translate(replacements)


def clean_text(v):
    if v is None:
        return None
    s = str(v).strip().replace('\ufffc', '').replace('\xa0', ' ')
    # Colapsar espacios multiples
    s = re.sub(r'\s+', ' ', s)
    # Escapar comillas simples para SQL
    s = s.replace("'", "''")
    return s if s else None


def normalize_doc_type(raw):
    if not raw:
        return 'TI'
    s = str(raw).strip().upper().replace(',', '.').replace(' ', '')
    if 'REGISTRO' in s or s == 'RC':
        return 'RC'
    if 'CEDULA' in s or 'CÉDULA' in s.upper() or s == 'CC':
        return 'CC'
    if 'TARJETA' in s.upper() or s.startswith('T') or s == 'TI':
        return 'TI'
    return 'TI'  # fallback


def normalize_doc_number(raw):
    if not raw:
        return None
    s = str(raw).strip()
    # Quitar decimal .0 de floats
    if s.endswith('.0'):
        s = s[:-2]
    # Quitar puntos, espacios, guiones
    s = re.sub(r'[^\d]', '', s)
    return s if s else None


def normalize_phone(raw):
    if not raw:
        return None
    s = str(raw).strip()
    if s.endswith('.0'):
        s = s[:-2]
    # Guardar primer telefono si hay varios separados por / o -
    parts = re.split(r'[/\-]', s)
    first = parts[0].strip()
    # Limpiar
    first = re.sub(r'[^\d]', '', first)
    # Validar largo (celular colombiano: 10 digitos, fijo: 7-10)
    if len(first) > 10:
        return first[:10]  # tomar primeros 10
    return first if first else None


def parse_dob(raw, sheet, name):
    # Override manual si aplica (normalizando ambos lados)
    name_norm = strip_accents(str(name or '').strip()).lower()
    sheet_norm = sheet.strip()
    for (s_key, n_key), d in DATE_OVERRIDES.items():
        if s_key == sheet_norm and strip_accents(n_key).lower() == name_norm:
            return d
    if raw is None:
        return None
    if isinstance(raw, datetime):
        return raw.date()
    if isinstance(raw, date):
        return raw
    # Intentar parsear string DD/MM/YYYY
    s = str(raw).strip()
    s = re.sub(r'^[^\d]+', '', s)  # Quitar prefijos tipo ":"
    m = re.match(r'^(\d{1,2})/(\d{1,2})/(\d{4})$', s)
    if m:
        d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
        try:
            return date(y, mo, d)
        except ValueError:
            return None
    return None


def sql_str(v):
    if v is None:
        return 'NULL'
    return f"'{v}'"


def sql_date(d):
    if d is None:
        return 'NULL'
    return f"'{d.isoformat()}'"


def build_emergency_contact(parent_name, phone):
    """Combina nombre y telefono del acudiente en un solo campo editable."""
    parts = []
    if parent_name:
        parts.append(parent_name)
    if phone:
        parts.append(f"Tel: {phone}")
    return ' | '.join(parts) if parts else None


def extract_team_data(wb, sheet_name):
    """Extrae y normaliza atletas de una hoja."""
    ws = wb[sheet_name]
    col_map = SHEET_COLUMN_MAP[sheet_name]
    team_clean = sheet_name.strip()
    athletes = []
    pending = []  # atletas con data incompleta

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        values = list(row)
        # Extraer por mapeo
        record = {}
        for col_idx, field in col_map.items():
            if col_idx < len(values):
                record[field] = values[col_idx]

        name_raw = record.get('name')
        if not name_raw or not str(name_raw).strip():
            continue  # fila vacia

        name = clean_text(name_raw)
        if not name:
            continue

        # Normalizar todos los campos
        doc_type = normalize_doc_type(record.get('doc_type'))
        doc_number = normalize_doc_number(record.get('doc_number'))
        dob = parse_dob(record.get('dob'), sheet_name, name_raw)
        tshirt = clean_text(record.get('tshirt'))
        eps = clean_text(record.get('eps'))
        phone_athlete = normalize_phone(record.get('phone_athlete'))
        parent_name = clean_text(record.get('parent_name'))
        phone_parent = normalize_phone(record.get('phone_parent'))

        # Fallback: si no hay phone_parent pero hay phone_athlete, se considera del acudiente
        primary_phone = phone_parent or phone_athlete

        athlete = {
            'team': team_clean,
            'row': row_idx,
            'full_name': name,
            'doc_type': doc_type,
            'doc_number': doc_number,
            'dob': dob,
            'tshirt': tshirt,
            'eps': eps,
            'parent_name': parent_name,
            'parent_phone': primary_phone,
            'athlete_phone': phone_athlete if phone_parent else None,  # solo si hay ambos
        }

        # Determinar si esta pendiente
        missing = []
        if not dob:
            missing.append('fecha_nacimiento')
        if not doc_number:
            missing.append('documento')
        if not parent_name and not primary_phone:
            missing.append('acudiente')
        if not eps:
            missing.append('eps')

        if missing:
            athlete['pending_reason'] = ','.join(missing)
            pending.append(athlete)

        athletes.append(athlete)

    return athletes, pending


def generate_sql(all_athletes, pending_athletes):
    lines = []
    lines.append("-- =========================================================================")
    lines.append("-- CARGA: 6 equipos y 102 atletas Fontibon")
    lines.append(f"-- Owner:   {OWNER_EMAIL}")
    lines.append(f"-- Escuela: Spirit Fontibon (Test) [{SCHOOL_ID}]")
    lines.append(f"-- Sede:    Fontibon              [{BRANCH_ID}]")
    lines.append("-- SQL plano (sin DO block, sin variables). Idempotente.")
    lines.append("-- =========================================================================")
    lines.append("")
    lines.append("-- ========================================")
    lines.append("-- 1. LIMPIEZA (idempotente): borra equipos demo previos y sus atletas")
    lines.append("-- ========================================")
    lines.append("DELETE FROM public.children")
    lines.append(f" WHERE school_id = '{SCHOOL_ID}'")
    lines.append("   AND is_demo = true")
    lines.append("   AND team_id IN (")
    lines.append("       SELECT id FROM public.teams")
    lines.append(f"        WHERE school_id = '{SCHOOL_ID}'")
    lines.append("          AND is_demo = true")
    lines.append("          AND name IN ('SPRINKLES','BUTTERFLY','BOMBSQUAD','LEGENDS','FIRESQUAD','BOMBSHELLS')")
    lines.append("   );")
    lines.append("")
    lines.append("DELETE FROM public.teams")
    lines.append(f" WHERE school_id = '{SCHOOL_ID}'")
    lines.append("   AND is_demo = true")
    lines.append("   AND name IN ('SPRINKLES','BUTTERFLY','BOMBSQUAD','LEGENDS','FIRESQUAD','BOMBSHELLS');")
    lines.append("")
    lines.append("-- ========================================")
    lines.append("-- 2. EQUIPOS (6)")
    lines.append("-- ========================================")
    lines.append("INSERT INTO public.teams (id, school_id, branch_id, name, sport, age_group, season, max_students, is_demo) VALUES")

    team_inserts = []
    for team_name, uuid_literal in TEAM_UUIDS.items():
        age_group = TEAM_AGE_GROUPS[team_name]
        team_inserts.append(
            f"    ({uuid_literal}, '{SCHOOL_ID}', '{BRANCH_ID}', "
            f"'{team_name}', 'Cheerleading All-Star', '{age_group}', '2026', 30, true)"
        )
    lines.append(',\n'.join(team_inserts) + ';')
    lines.append("")
    lines.append("-- ========================================")
    lines.append("-- 3. ATLETAS (children)")
    lines.append("-- ========================================")
    lines.append("INSERT INTO public.children (")
    lines.append("    school_id, branch_id, team_id, parent_id,")
    lines.append("    full_name, date_of_birth, doc_type, doc_number,")
    lines.append("    emergency_contact, medical_info, monthly_fee, is_demo")
    lines.append(") VALUES")

    athlete_rows = []
    pending_set = {(a['team'], a['full_name']) for a in pending_athletes}
    PLACEHOLDER_DOB = date(2020, 1, 1)

    for a in all_athletes:
        team_uuid = TEAM_UUIDS[a['team']]
        emergency = build_emergency_contact(a['parent_name'], a['parent_phone'])
        medical_parts = []
        if a['eps']:
            medical_parts.append(f"EPS: {a['eps']}")
        if a['tshirt']:
            medical_parts.append(f"Talla: {a['tshirt']}")
        if a['athlete_phone']:
            medical_parts.append(f"Tel atleta: {a['athlete_phone']}")
        is_pending = (a['team'], a['full_name']) in pending_set
        dob_to_use = a['dob']
        if dob_to_use is None:
            dob_to_use = PLACEHOLDER_DOB
            medical_parts.insert(0, "*** FECHA PLACEHOLDER - CORREGIR ***")
        if is_pending:
            medical_parts.append(f"[PENDIENTE COMPLETAR: {a.get('pending_reason','')}]")
        medical = ' | '.join(medical_parts) if medical_parts else None

        row_sql = (
            f"    ('{SCHOOL_ID}', '{BRANCH_ID}', {team_uuid}, NULL, "
            f"{sql_str(a['full_name'])}, {sql_date(dob_to_use)}, "
            f"{sql_str(a['doc_type'])}, {sql_str(a['doc_number'])}, "
            f"{sql_str(emergency)}, {sql_str(medical)}, 0, true)"
        )
        athlete_rows.append(row_sql)

    lines.append(',\n'.join(athlete_rows) + ';')
    lines.append("")
    lines.append("-- ========================================")
    lines.append("-- 4. VERIFICACION RAPIDA")
    lines.append("-- ========================================")
    lines.append(f"SELECT 'resultado' AS tipo, COUNT(*) AS equipos FROM public.teams WHERE school_id = '{SCHOOL_ID}' AND is_demo = true AND name IN ('SPRINKLES','BUTTERFLY','BOMBSQUAD','LEGENDS','FIRESQUAD','BOMBSHELLS');")
    lines.append(f"SELECT 'resultado' AS tipo, COUNT(*) AS atletas FROM public.children WHERE school_id = '{SCHOOL_ID}' AND is_demo = true AND team_id IN (SELECT id FROM public.teams WHERE school_id = '{SCHOOL_ID}' AND is_demo = true AND name IN ('SPRINKLES','BUTTERFLY','BOMBSQUAD','LEGENDS','FIRESQUAD','BOMBSHELLS'));")
    lines.append("")
    lines.append("-- =========================================================================")
    lines.append("-- ATLETAS MARCADOS COMO PENDIENTES (info incompleta en Excel)")
    lines.append("-- Editar en la app: school dashboard > atleta > completar datos")
    lines.append("-- =========================================================================")
    for a in pending_athletes:
        lines.append(f"-- [{a['team']}] {a['full_name']} -- falta: {a.get('pending_reason','')}")

    return '\n'.join(lines)


def main():
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    all_athletes = []
    all_pending = []

    print(f"{'='*70}")
    print(f"EXTRACCION Y NORMALIZACION FONTIBON 2026")
    print(f"{'='*70}\n")

    for sheet in wb.sheetnames:
        if sheet.strip() == 'Hoja 2' or sheet.strip() not in [k.strip() for k in SHEET_COLUMN_MAP.keys()]:
            continue
        # Usar la clave exacta del map
        map_key = next(k for k in SHEET_COLUMN_MAP.keys() if k.strip() == sheet.strip())
        athletes, pending = extract_team_data(wb, sheet)
        # Usar nombre limpio del equipo
        for a in athletes:
            a['team'] = sheet.strip()
        for p in pending:
            p['team'] = sheet.strip()
        all_athletes.extend(athletes)
        all_pending.extend(pending)
        print(f"  {sheet.strip():12s} -> {len(athletes):3d} atletas ({len(pending)} pendientes)")

    print(f"\n  {'TOTAL':12s} -> {len(all_athletes)} atletas ({len(all_pending)} pendientes)\n")

    # Mostrar pendientes
    if all_pending:
        print(f"{'='*70}")
        print(f"ATLETAS PENDIENTES DE COMPLETAR (incluidos pero con info faltante)")
        print(f"{'='*70}")
        for a in all_pending:
            print(f"  [{a['team']:12s}] {a['full_name']:45s} falta: {a.get('pending_reason','')}")
        print()

    # Generar SQL
    sql = generate_sql(all_athletes, all_pending)
    with open(OUTPUT_SQL, 'w', encoding='utf-8') as f:
        f.write(sql)

    print(f"{'='*70}")
    print(f"SQL GENERADO EN:")
    print(f"  {OUTPUT_SQL}")
    print(f"{'='*70}")
    print(f"  Lineas: {len(sql.splitlines())}")
    print(f"  Tamaño: {len(sql)/1024:.1f} KB")


if __name__ == '__main__':
    main()
