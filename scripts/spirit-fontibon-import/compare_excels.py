"""
Compara FONTIBON 2026.xlsx (original) vs FONTIBON 2026 (1).xlsx (actualizado).
Usa HEADERS por nombre (no por posicion) — el Excel nuevo agrego columna RH.
Genera SQL UPDATE solo con diferencias, match por doc_number.
"""
import openpyxl, sys, io, re
from datetime import date, datetime

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

OLD_XLSX = 'C:/Users/Usuario/Documents/DOCUMENTACION SPORTMPAS/FONTIBON 2026.xlsx'
NEW_XLSX = 'C:/Users/Usuario/Documents/DOCUMENTACION SPORTMPAS/FONTIBON 2026 (1).xlsx'
OUT_SQL  = 'c:/Users/Usuario/Documents/demo/sportmaps-demo/scripts/spirit-fontibon-import/09_apply_excel_updates.sql'
SCHOOL_ID = 'aaaaaaaa-bbbb-cccc-dddd-eeeeeeeeeeee'

# Aliases de headers a campo interno (todos lowercase, sin acentos)
HEADER_ALIASES = {
    'name':          ['nombre de deportista', 'nombre'],
    'doc_type':      ['tipo de documento', 't.documento', 't,documento'],
    'doc_number':    ['no de documento', 'no documento'],
    'dob':           ['fecha de nacimiento'],
    'age':           ['edad'],
    'tshirt':        ['talla de camiseta', 't shirt', 'tshirt'],
    'eps':           ['eps'],
    'rh':            ['rh'],
    'phone_athlete': ['telefono deportista'],
    'parent_name':   ['nombre de acudiente'],
    'phone_parent':  ['telefono de acuente', 'telefono de acudiente', 'telefono'],
}


def norm_header(h):
    if not h: return ''
    s = str(h).strip().lower()
    s = s.replace('á','a').replace('é','e').replace('í','i').replace('ó','o').replace('ú','u').replace('ñ','n')
    s = re.sub(r'\s+', ' ', s)
    return s


def build_col_map(ws):
    """Mapea cada field interno a su indice de columna usando los headers."""
    headers = [norm_header(c.value) for c in ws[1]]
    col_map = {}
    for field, aliases in HEADER_ALIASES.items():
        for alias in aliases:
            alias_n = norm_header(alias)
            if alias_n in headers:
                col_map[field] = headers.index(alias_n)
                break
    return col_map


def clean(v):
    if v is None: return None
    s = str(v).strip().replace('\ufffc', '').replace('\xa0', ' ')
    s = re.sub(r'\s+', ' ', s)
    # Normalizar flotantes que representan enteros: "10.0" -> "10"
    if re.match(r'^\d+\.0$', s):
        s = s[:-2]
    return s if s else None


def norm_compare(v):
    """Normaliza para comparar: case-insensitive, sin espacios extra."""
    if v is None: return ''
    return str(v).strip().lower()


def norm_doc(v):
    if v is None: return None
    s = str(v).strip()
    if s.endswith('.0'): s = s[:-2]
    s = re.sub(r'[^\d]', '', s)
    return s if s else None


def norm_phone(v):
    if v is None: return None
    s = str(v).strip()
    if s.endswith('.0'): s = s[:-2]
    parts = re.split(r'[/\-]', s)
    first = re.sub(r'[^\d]', '', parts[0].strip())
    return first[:10] if len(first) > 10 else (first if first else None)


def parse_dob(v):
    if v is None: return None
    if isinstance(v, datetime): return v.date()
    if isinstance(v, date):     return v
    s = str(v).strip()
    s = re.sub(r'^[^\d]+', '', s)
    m = re.match(r'^(\d{1,2})/(\d{1,2})/(\d{4})$', s)
    if m:
        d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
        try: return date(y, mo, d)
        except ValueError: return None
    return None


def extract_athletes(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    out = []
    for sheet in wb.sheetnames:
        if sheet.strip() == 'Hoja 2': continue
        ws = wb[sheet]
        col_map = build_col_map(ws)
        if 'name' not in col_map: continue  # hoja sin headers validos

        def get(row, field):
            i = col_map.get(field)
            return row[i] if (i is not None and i < len(row)) else None

        for row in ws.iter_rows(min_row=2, values_only=True):
            name = clean(get(row, 'name'))
            if not name: continue
            out.append({
                'team':          sheet.strip(),
                'name':          name,
                'doc_number':    norm_doc(get(row, 'doc_number')),
                'dob':           parse_dob(get(row, 'dob')),
                'eps':           clean(get(row, 'eps')),
                'tshirt':        clean(get(row, 'tshirt')),
                'rh':            clean(get(row, 'rh')),
                'parent_name':   clean(get(row, 'parent_name')),
                'parent_phone':  norm_phone(get(row, 'phone_parent')),
            })
    return out


def sql_str(v):
    if v is None: return 'NULL'
    return "'" + str(v).replace("'", "''") + "'"


def main():
    print(f'Leyendo OLD: {OLD_XLSX}')
    old_list = extract_athletes(OLD_XLSX)
    print(f'Leyendo NEW: {NEW_XLSX}')
    new_list = extract_athletes(NEW_XLSX)

    old_by_doc = {a['doc_number']: a for a in old_list if a['doc_number']}
    new_by_doc = {a['doc_number']: a for a in new_list if a['doc_number']}

    print(f'\nOLD atletas: {len(old_list)}  |  NEW atletas: {len(new_list)}')
    print(f'OLD con doc: {len(old_by_doc)}   |  NEW con doc: {len(new_by_doc)}')

    diffs = []
    for doc, new_a in new_by_doc.items():
        old_a = old_by_doc.get(doc)
        if old_a is None:
            diffs.append(('DOC_NUEVO', new_a, {}))
            continue
        changes = {}
        for field in ['name', 'dob', 'eps', 'tshirt', 'rh', 'parent_name', 'parent_phone']:
            ov = old_a.get(field)
            nv = new_a.get(field)
            # Dates se comparan directamente. Strings normalizados.
            if isinstance(ov, date) or isinstance(nv, date):
                if ov != nv:
                    changes[field] = (ov, nv)
            else:
                if norm_compare(ov) != norm_compare(nv):
                    changes[field] = (ov, nv)
        if changes:
            diffs.append(('CAMBIO', new_a, changes))

    print(f'\n{"="*90}')
    print(f'DIFERENCIAS ENCONTRADAS: {len(diffs)}')
    print(f'{"="*90}\n')

    for tipo, athlete, changes in diffs[:30]:  # primer 30 para no spamear
        print(f'[{tipo}] {athlete["team"]:12s} | {athlete["name"]}')
        for field, (old_v, new_v) in changes.items():
            print(f'    {field:14s}: {old_v!s:40s} -> {new_v!s}')
        print()

    if len(diffs) > 30:
        print(f'... y {len(diffs) - 30} mas (ver SQL generado)\n')

    # Generar SQL UPDATE
    lines = []
    lines.append('-- =========================================================================')
    lines.append('-- Diferencias FONTIBON 2026.xlsx -> FONTIBON 2026 (1).xlsx')
    lines.append(f'-- Total cambios: {len(diffs)}')
    lines.append('-- Match por doc_number (match seguro)')
    lines.append('-- =========================================================================')
    lines.append('')

    for tipo, a, changes in diffs:
        if not changes: continue
        if not a['doc_number']: continue
        set_parts = []
        if 'name' in changes:
            set_parts.append(f"full_name = {sql_str(a['name'])}")
        if 'dob' in changes and a['dob']:
            set_parts.append(f"date_of_birth = '{a['dob'].isoformat()}'")
        if 'parent_name' in changes or 'parent_phone' in changes:
            ec_parts = []
            if a['parent_name']: ec_parts.append(a['parent_name'])
            if a['parent_phone']: ec_parts.append(f"Tel: {a['parent_phone']}")
            ec = ' | '.join(ec_parts) if ec_parts else None
            set_parts.append(f"emergency_contact = {sql_str(ec)}")
        if 'eps' in changes or 'tshirt' in changes or 'rh' in changes:
            m_parts = []
            if a['eps']:    m_parts.append(f"EPS: {a['eps']}")
            if a['tshirt']: m_parts.append(f"Talla: {a['tshirt']}")
            if a['rh']:     m_parts.append(f"RH: {a['rh']}")
            m = ' | '.join(m_parts) if m_parts else None
            set_parts.append(f"medical_info = {sql_str(m)}")
        if not set_parts: continue
        set_parts.append('updated_at = now()')

        lines.append(f"-- [{a['team']}] {a['name']}")
        lines.append(f"UPDATE public.children SET {', '.join(set_parts)}")
        lines.append(f" WHERE school_id = '{SCHOOL_ID}' AND doc_number = {sql_str(a['doc_number'])};")
        lines.append('')

    # Verificacion final
    lines.append('-- Verificacion: cuantos atletas tienen RH ahora')
    lines.append(f"SELECT COUNT(*) FROM public.children WHERE school_id = '{SCHOOL_ID}' AND medical_info LIKE '%RH:%';")

    with open(OUT_SQL, 'w', encoding='utf-8') as f:
        f.write('\n'.join(lines))

    print(f'\nSQL generado en: {OUT_SQL}')


if __name__ == '__main__':
    main()
