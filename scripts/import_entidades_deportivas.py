#!/usr/bin/env python3
"""
import_entidades_deportivas.py
===============================

Procesa los 3 XLSX oficiales de Mindeporte/Coldeportes en C:\\Users\\Usuario\\Documents\\esxcuelas:

  1. Directorio-Institutos-Departamentales-y-Municipales-del-Deporte-2026.xlsx
     - Hoja "Departamentales" (~27 entes departamentales)
     - Hoja "Municipales_Distritales" (~38 entes municipales)
  2. Directorio-Federaciones-Deportivas-2025.xlsx (79 federaciones con direccion)
  3. Directorio-Asociaciones-Recreativas-2025.xlsx (~10 asociaciones)

Para cada uno:
  - Normaliza encoding (CP1252 -> UTF8).
  - Para Federaciones/Asociaciones: geocode con direccion + domicilio.
  - Para Institutos: geocode centroide de la ciudad/departamento.

Outputs (idempotentes, UPSERT por external_school_imports):
  supabase/seed/entidades_deportivas_2025_2026.sql
  Landing_page/.../mapData.entidades.ts
"""

from __future__ import annotations

import hashlib
import json
import os
import re
import sys
import time
import unicodedata
from pathlib import Path
from typing import Optional


def _hash8(s: str) -> str:
    """Hash MD5 truncado a 8 chars — unicidad garantizada de external_ref."""
    return hashlib.md5(s.encode("utf-8")).hexdigest()[:8]

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")  # type: ignore
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")  # type: ignore
except Exception:
    pass

try:
    import openpyxl  # type: ignore
    import requests  # type: ignore
except ImportError as e:
    print(f"Missing dep: {e.name}. Run: pip install openpyxl requests")
    sys.exit(1)


ROOT = Path(__file__).resolve().parents[1]
CACHE_FILE = ROOT / "scripts" / ".geocode_cache.json"
SQL_OUT = ROOT / "supabase" / "seed" / "entidades_deportivas_2025_2026.sql"
TS_OUT = Path(
    "C:/Users/Usuario/Documents/Landing_page/sportmap-maps-landing-page/frontend/src/data/mapData.entidades.ts"
)

XLSX_DIR = Path("C:/Users/Usuario/Documents/esxcuelas")
FILE_INSTITUTOS = XLSX_DIR / "Directorio-Institutos-Departamentales-y-Municipales-del-Deporte-2026.xlsx"
FILE_FEDERACIONES = XLSX_DIR / "Directorio-Federaciones-Deportivas-2025.xlsx"
FILE_ASOCIACIONES = XLSX_DIR / "Directorio-Asociaciones-Recreativas-2025.xlsx"

NOMINATIM_URL = "https://nominatim.openstreetmap.org/search"
RATE_LIMIT = 1.1
USER_AGENT = "SportMaps-Importer/1.0 (brayan.lopez@osigu.com)"

# Colombia bounding box
COLOMBIA_BOUNDS = (-4.5, 13.5, -82.0, -66.0)


# ── Normalizacion encoding (XLSX vienen en CP1252) ───────────────────────────
ENCODING_FIXES = {
    "�": "",  # no podemos adivinar — quitar
    "DEPARTAMENT�ALES": "DEPARTAMENTALES",
    "ACR�NIMO": "ACRÓNIMO",
    "RECREACI�N": "RECREACIÓN",
    "ATL�NTICO": "ATLÁNTICO",
    "BOL�VAR": "BOLÍVAR",
    "PE�A": "PEÑA",
    "MU�OZ": "MUÑOZ",
    "Bogot�": "Bogotá",
    "Medell�n": "Medellín",
    "Bel�n": "Belén",
    "B�rbara": "Bárbara",
}


def normalize(s) -> str:
    if s is None:
        return ""
    txt = str(s)
    for bad, good in ENCODING_FIXES.items():
        txt = txt.replace(bad, good)
    return re.sub(r"\s+", " ", txt).strip()


def slugify(s: str) -> str:
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode("ascii")
    s = re.sub(r"[^a-zA-Z0-9\s-]", "", s).lower()
    s = re.sub(r"\s+", "-", s).strip("-")
    return s[:60] or "ente"


def sql_str(s) -> str:
    if s is None or s == "":
        return "NULL"
    return "'" + str(s).replace("'", "''") + "'"


def sql_array(xs: list[str]) -> str:
    if not xs:
        return "ARRAY[]::text[]"
    return "ARRAY[" + ",".join(sql_str(x) for x in xs) + "]::text[]"


# ── Geocode cache ────────────────────────────────────────────────────────────
def load_cache() -> dict[str, dict]:
    if CACHE_FILE.exists():
        try:
            return json.loads(CACHE_FILE.read_text(encoding="utf-8"))
        except Exception:
            return {}
    return {}


def save_cache(cache: dict[str, dict]) -> None:
    CACHE_FILE.parent.mkdir(parents=True, exist_ok=True)
    CACHE_FILE.write_text(json.dumps(cache, ensure_ascii=False, indent=2), encoding="utf-8")


def in_colombia(lat: float, lng: float) -> bool:
    return COLOMBIA_BOUNDS[0] <= lat <= COLOMBIA_BOUNDS[1] and COLOMBIA_BOUNDS[2] <= lng <= COLOMBIA_BOUNDS[3]


def geocode(query: str, cache: dict[str, dict]) -> Optional[tuple[float, float]]:
    key = query.strip().lower()
    if not key:
        return None
    if key in cache:
        v = cache[key]
        return (float(v["lat"]), float(v["lng"])) if v.get("lat") is not None else None

    for attempt in range(3):
        try:
            time.sleep(RATE_LIMIT if attempt == 0 else 3.0 * attempt)
            resp = requests.get(
                NOMINATIM_URL,
                params={"q": query, "format": "json", "limit": 1, "countrycodes": "co"},
                headers={"User-Agent": USER_AGENT},
                timeout=30,
            )
            resp.raise_for_status()
            data = resp.json()
            if not data:
                cache[key] = {"lat": None, "lng": None}
                save_cache(cache)
                return None
            lat, lng = float(data[0]["lat"]), float(data[0]["lon"])
            if not in_colombia(lat, lng):
                cache[key] = {"lat": None, "lng": None}
                save_cache(cache)
                return None
            cache[key] = {"lat": lat, "lng": lng}
            save_cache(cache)
            return lat, lng
        except Exception as e:
            if attempt < 2:
                print(f"  [warn] geocode retry {attempt+1}: {str(e)[:80]}", flush=True)
                continue
    cache[key] = {"lat": None, "lng": None}
    save_cache(cache)
    return None


# ── Parsers por archivo ──────────────────────────────────────────────────────

def parse_institutos() -> list[dict]:
    """Lee ambas hojas (Departamentales + Municipales). Headers en row 2."""
    out: list[dict] = []
    if not FILE_INSTITUTOS.exists():
        print(f"[warn] {FILE_INSTITUTOS} no existe — skip")
        return out
    wb = openpyxl.load_workbook(FILE_INSTITUTOS, data_only=True)

    for sheet_name in ["Departamentales", "Municipales_Distritales"]:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        # Headers en row 2, data desde row 3
        for i, row in enumerate(ws.iter_rows(min_row=3, values_only=True), 3):
            if not row[2]:  # Ente vacio
                continue
            nivel = normalize(row[0])
            depto_ciudad = normalize(row[1])
            ente = normalize(row[2])
            acronimo = normalize(row[3])
            cargo = normalize(row[4])
            titular = normalize(row[5])
            telefonos = normalize(row[6])
            correo = normalize(row[7])

            out.append({
                "kind": "instituto",
                "ext_ref": f"INST-{sheet_name[:3].upper()}-{slugify(ente)[:30]}-{_hash8(ente)}",
                "name": ente,
                "acronym": acronimo,
                "city": depto_ciudad,
                "level": nivel,
                "description": f"{cargo}: {titular}. Ente {nivel.lower()} de deporte. {depto_ciudad}.",
                "phone": (telefonos.split("\n")[0].split("/")[0].strip() if telefonos else None),
                "email": correo if "@" in correo else None,
                "address": None,
                "sport": "Multideporte",
            })
    return out


def parse_federaciones() -> list[dict]:
    out: list[dict] = []
    if not FILE_FEDERACIONES.exists():
        print(f"[warn] {FILE_FEDERACIONES} no existe — skip")
        return out
    wb = openpyxl.load_workbook(FILE_FEDERACIONES, data_only=True)
    ws = wb["Mapa 2025"]
    # Headers row 3, data row 4+
    for i, row in enumerate(ws.iter_rows(min_row=4, values_only=True), 4):
        if not row[2]:
            continue
        no = row[0]
        nit = normalize(row[1])
        nombre = normalize(row[2])
        rep_legal = normalize(row[3])
        domicilio = normalize(row[5])
        direccion = normalize(row[6])
        correo = normalize(row[7])
        if "@" in correo:
            correo = correo.split(";")[0].strip()
        else:
            correo = None
        telefono = normalize(row[8])
        telefono_clean = re.findall(r"\d{7,}", telefono.replace("\n", " "))
        phone_first = telefono_clean[0] if telefono_clean else None

        # Extraer deporte del nombre: 'FEDERACION COLOMBIANA DE <DEPORTE>'
        sport_match = re.search(r"FEDERACION (?:COLOMBIANA )?(?:DE )?(.+)", nombre, re.IGNORECASE)
        sport = sport_match.group(1).strip() if sport_match else "Multideporte"
        # Limpiar "ACTIVIDADES" etc.
        sport = re.sub(r"^(ACTIVIDADES |ARQUEROS DE COLOMBIA|AUTOMOVILISMO DEPORTIVO)$", lambda m: m.group(0), sport, flags=re.IGNORECASE)

        out.append({
            "kind": "federacion",
            "ext_ref": f"FED-{slugify(nombre)[:40]}-{_hash8(nombre)}",
            "name": nombre.title(),
            "acronym": None,
            "city": domicilio,
            "address": direccion,
            "phone": phone_first,
            "email": correo,
            "description": f"Federación deportiva colombiana. Representante: {rep_legal.title()}. NIT: {nit}.",
            "sport": sport.title(),
        })
    return out


def parse_asociaciones() -> list[dict]:
    out: list[dict] = []
    if not FILE_ASOCIACIONES.exists():
        print(f"[warn] {FILE_ASOCIACIONES} no existe — skip")
        return out
    wb = openpyxl.load_workbook(FILE_ASOCIACIONES, data_only=True)
    ws = wb["Hoja1"]
    # Headers row 7, data row 8+. Cols: B=No, C=Asociacion, D=Rep, E=City, F=Dir, G=Email, H=Tel
    for i, row in enumerate(ws.iter_rows(min_row=8, values_only=True), 8):
        if not row[2]:  # col C
            continue
        nombre = normalize(row[2])
        rep_legal = normalize(row[3])
        ciudad = normalize(row[4])
        direccion = normalize(row[5])
        correo = normalize(row[6]) if row[6] and "@" in str(row[6]) else None
        telefono = normalize(row[7]) if row[7] else None
        telefono_clean = re.findall(r"\d{7,}", str(telefono or "").replace("\n", " "))
        phone_first = telefono_clean[0] if telefono_clean else None

        sport_match = re.search(r"ASOCIACI[OÓ]N (?:COLOMBIANA )?(?:DE )?(.+)", nombre, re.IGNORECASE)
        sport = sport_match.group(1).strip() if sport_match else "Multideporte"

        out.append({
            "kind": "asociacion",
            "ext_ref": f"ASOC-{slugify(nombre)[:40]}-{_hash8(nombre)}",
            "name": nombre.title(),
            "acronym": None,
            "city": ciudad,
            "address": direccion,
            "phone": phone_first,
            "email": correo,
            "description": f"Asociación deportiva recreativa. Representante: {rep_legal.title()}.",
            "sport": sport.title(),
        })
    return out


# ── SQL emission ─────────────────────────────────────────────────────────────

def write_sql(records: list[dict]) -> None:
    SQL_OUT.parent.mkdir(parents=True, exist_ok=True)
    lines = [
        "-- ============================================================",
        "-- SPORTMAPS — Entidades deportivas oficiales 2025/2026",
        "-- Fuente: Directorios oficiales Mindeporte/Coldeportes",
        "-- (Institutos Dept/Mun 2026 + Federaciones 2025 + Asociaciones 2025)",
        "-- Generado por scripts/import_entidades_deportivas.py",
        "-- ============================================================",
        "",
        "BEGIN;",
        "",
    ]

    # Mapeo kind -> school_type (ya no son todos 'academy')
    KIND_TO_TYPE = {
        "instituto":  "institute",
        "federacion": "federation",
        "asociacion": "association",
    }

    for rec in records:
        slug = slugify(rec["name"]) + "-" + _hash8(rec["ext_ref"])
        sports = [rec["sport"]] if rec.get("sport") else ["Multideporte"]
        school_type = KIND_TO_TYPE.get(rec["kind"], "academy")
        raw = json.dumps({
            "kind": rec["kind"],
            "acronym": rec.get("acronym"),
            "level": rec.get("level"),
        }, ensure_ascii=False)

        lat_sql = f"{rec['lat']:.7f}" if rec.get("lat") else "NULL"
        lng_sql = f"{rec['lng']:.7f}" if rec.get("lng") else "NULL"

        lines.append(f"-- {rec['name']} ({rec['ext_ref']})")
        lines.append("DO $$")
        lines.append("DECLARE v_school_id uuid; v_existing uuid;")
        lines.append("BEGIN")
        lines.append(f"  SELECT school_id INTO v_existing FROM public.external_school_imports WHERE external_ref = {sql_str(rec['ext_ref'])};")
        lines.append("  IF v_existing IS NULL THEN")
        lines.append("    INSERT INTO public.schools (")
        lines.append("      name, description, school_type, city, address, phone, email, sports,")
        lines.append("      verified, is_demo, slug, onboarding_status")
        lines.append("    ) VALUES (")
        lines.append(f"      {sql_str(rec['name'])},")
        lines.append(f"      {sql_str(rec['description'])},")
        lines.append(f"      '{school_type}',")
        lines.append(f"      {sql_str(rec.get('city') or 'Colombia')},")
        lines.append(f"      {sql_str(rec.get('address'))},")
        lines.append(f"      {sql_str(rec.get('phone'))},")
        lines.append(f"      {sql_str(rec.get('email'))},")
        lines.append(f"      {sql_array(sports)},")
        lines.append("      true, false,")
        lines.append(f"      {sql_str(slug)},")
        lines.append("      'completed'")
        lines.append("    ) RETURNING id INTO v_school_id;")
        lines.append("")
        lines.append("    INSERT INTO public.external_school_imports (source, external_ref, school_id, raw_payload)")
        lines.append(f"    VALUES ('mindeporte_entidades_2025_2026', {sql_str(rec['ext_ref'])}, v_school_id, {sql_str(raw)}::jsonb);")
        lines.append("  ELSE")
        lines.append("    v_school_id := v_existing;")
        lines.append("    UPDATE public.schools SET")
        lines.append(f"      description = COALESCE({sql_str(rec['description'])}, description),")
        lines.append(f"      phone       = COALESCE({sql_str(rec.get('phone'))}, phone),")
        lines.append(f"      email       = COALESCE({sql_str(rec.get('email'))}, email),")
        lines.append("      verified    = true,")
        lines.append("      updated_at  = now()")
        lines.append("    WHERE id = v_school_id;")
        lines.append("  END IF;")
        lines.append("")
        lines.append("  INSERT INTO public.school_settings (school_id) VALUES (v_school_id) ON CONFLICT (school_id) DO NOTHING;")
        lines.append("  UPDATE public.school_settings SET public_profile_enabled = true WHERE school_id = v_school_id;")
        lines.append("")
        if rec.get("lat") is not None:
            lines.append("  INSERT INTO public.school_branches (school_id, name, address, city, phone, lat, lng, is_main, status)")
            lines.append("  SELECT v_school_id, 'Sede Principal',")
            lines.append(f"         {sql_str(rec.get('address'))}, {sql_str(rec.get('city') or 'Colombia')},")
            lines.append(f"         {sql_str(rec.get('phone'))}, {lat_sql}, {lng_sql}, true, 'active'")
            lines.append("  WHERE NOT EXISTS (SELECT 1 FROM public.school_branches WHERE school_id = v_school_id AND is_main = true);")
        lines.append("END $$;")
        lines.append("")

    lines.append("COMMIT;")
    SQL_OUT.write_text("\n".join(lines), encoding="utf-8")


def write_ts(records: list[dict]) -> None:
    TS_OUT.parent.mkdir(parents=True, exist_ok=True)
    lines = [
        "// Auto-generado por scripts/import_entidades_deportivas.py — NO EDITAR A MANO",
        "// Fuente: Directorios Mindeporte (Institutos + Federaciones + Asociaciones)",
        "",
        "import type { MapLocation } from './mapData';",
        "",
        "export const entidadesDeportivasOficiales: MapLocation[] = [",
    ]
    # Mapeo kind -> entityType del landing (para filtros UI)
    KIND_TO_ENTITY_TYPE = {
        "instituto":  "institute",
        "federacion": "federation",
        "asociacion": "association",
    }
    for rec in records:
        if not rec.get("lat") or not rec.get("lng"):
            continue
        name = (rec["name"] or "").replace("'", "\\'")
        sport = (rec.get("sport") or "Multideporte").replace("'", "\\'")
        city = (rec.get("city") or "Colombia").replace("'", "\\'")
        desc = (rec.get("description") or "").replace("'", "\\'")[:180]
        addr = (rec.get("address") or "").replace("'", "\\'")
        entity_type = KIND_TO_ENTITY_TYPE.get(rec["kind"], "academy")
        lines.append("  {")
        lines.append(f"    id: '{rec['ext_ref'].lower()}',")
        lines.append(f"    name: '{name}',")
        lines.append("    type: 'academy',")
        lines.append(f"    entityType: '{entity_type}',")
        lines.append(f"    sport: '{sport}',")
        lines.append(f"    lat: {rec['lat']:.7f},")
        lines.append(f"    lng: {rec['lng']:.7f},")
        lines.append(f"    city: '{city}',")
        lines.append(f"    description: '{desc}',")
        if addr:
            lines.append(f"    address: '{addr}',")
        if rec.get("phone"):
            lines.append(f"    phone: '+57 {rec['phone']}',")
        lines.append("  },")
    lines.append("];")
    TS_OUT.write_text("\n".join(lines), encoding="utf-8")


# ── Main ─────────────────────────────────────────────────────────────────────

def main() -> None:
    cache = load_cache()
    print(f"Cache geocode: {len(cache)} entries")

    print("\n=== Parseando archivos ===")
    institutos = parse_institutos()
    print(f"  Institutos: {len(institutos)}")
    federaciones = parse_federaciones()
    print(f"  Federaciones: {len(federaciones)}")
    asociaciones = parse_asociaciones()
    print(f"  Asociaciones: {len(asociaciones)}")

    all_records = institutos + federaciones + asociaciones
    print(f"\nTotal entidades: {len(all_records)}")

    print("\n=== Geocodificando ===")
    for idx, rec in enumerate(all_records, 1):
        candidates = []
        if rec.get("address") and rec.get("city"):
            candidates.append(f"{rec['address']}, {rec['city']}, Colombia")
        if rec.get("city"):
            candidates.append(f"{rec['city']}, Colombia")

        lat = lng = None
        for q in candidates:
            coords = geocode(q, cache)
            if coords:
                lat, lng = coords
                break
        rec["lat"] = lat
        rec["lng"] = lng
        marker = f"{lat:.4f},{lng:.4f}" if lat else "NO_GEO"
        print(f"  [{idx:03d}/{len(all_records)}] {rec['kind'][:4]} {rec['name'][:50]:50}  {marker}", flush=True)

    geocoded = sum(1 for r in all_records if r.get("lat"))
    print(f"\nGeocodificados: {geocoded}/{len(all_records)}")

    write_sql(all_records)
    write_ts(all_records)
    print(f"\n[ok] SQL: {SQL_OUT}")
    print(f"[ok] TS:  {TS_OUT}")


if __name__ == "__main__":
    main()
